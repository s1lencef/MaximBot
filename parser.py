import openpyxl

from models import ArtistModel, Statistics



def process(path, artist_name=None):
    obj = openpyxl.load_workbook(path)

    sheet = obj.active

    row = sheet.max_row
    answ = "Обработка завершена успешно!"

    for i in range(1, row + 1):
        name_cell = sheet.cell(row=i, column=1).value
        year_cell = sheet.cell(row=i, column=2).value
        first = sheet.cell(row=i, column=3).value

        name = artist_name if artist_name else name_cell
        print(name)
        try:
            artist = ArtistModel.get(ArtistModel.name == name)
        except Exception:
            return "Такого артиста не существует"

        print(artist)
        for j in range(1, 5):
            q1_cell = sheet.cell(row=i, column=j + 2).value
            if q1_cell == None:
                q1_cell = first
            print(q1_cell)
            staistics = Statistics(artist_id=artist, year=year_cell, quarter=j, state=q1_cell)
            try:
                staistics.save()
            except Exception:
                answ = "Не все строки документа были обработаны, проверьте структуру документа!"

    return answ